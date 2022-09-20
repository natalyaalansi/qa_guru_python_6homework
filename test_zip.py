from zipfile import ZipFile
from PyPDF2 import PdfReader
import csv
from openpyxl import load_workbook

filenames = ['sample1.xlsx', 'sample2.csv', 'sample3.pdf']

with ZipFile('resources/sample.zip', 'w') as archive:
   for filename in filenames:
      archive.write(filename)

with ZipFile('resources/sample.zip', 'r') as archive:
   archive.extractall('resources')

def test_xlsx():
   workbook = load_workbook('resources/sample1.xlsx')
   sheet = workbook.active
   assert sheet.cell(row=3, column=5).value == 'Great Britain'
   for x in range(1, 10):
      for y in range(1, 10):
         print(sheet.cell(row=x, column=y).value)

def test_csv():
   with open('resources/sample2.csv') as csvfile:
      table = csv.reader(csvfile)
      row_count = sum(1 for row in table)
      assert row_count == 10
      for line_no, line in enumerate(table, 1):
         if line_no == 2:
            print(line[1])

def test_pdf():
   pdf_reader = PdfReader('resources/sample3.pdf')
   number_of_pages = len(pdf_reader.pages)
   assert number_of_pages == 1
   page = pdf_reader.pages[0]
   text = page.extract_text()
   print(text)
   assert 'PDF' in text
